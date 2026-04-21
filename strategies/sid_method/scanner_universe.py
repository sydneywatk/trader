"""Universe scanner — backtest ~500 tickers and rank by win rate.

Creates two Excel files:
  1. sid_universe_ranked_YYYYMMDD.xlsx  — all qualifying tickers
  2. sid_top100_watchlist_YYYYMMDD.xlsx — top 100 + summary sheet

SPY filter rule change: SPY AND gate applies to SHORTS only.
Longs pass unconditionally (monkeypatched below).
"""

import json
import os
import sys
import tempfile
import time
from collections import Counter, defaultdict
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))  # trader/

# ── Imports from existing codebase ──────────────────────────────────────────
from config import CACHE_DIR, OUTPUT_DIR

# Override: config.MIN_QUALIFYING_TRADES is 15 but most tickers produce
# 8-14 trades over 2020-2025. Use 10 for this universe scan.
MIN_QUALIFYING_TRADES = 10
from shared.data import fetch_daily, fetch_weekly
from shared.indicators import add_daily_indicators, add_weekly_rsi
from signals import find_rsi_signals
from shared.earnings import fetch_earnings_dates
import backtest
from backtest import run_backtest_for_ticker
from universe import get_universe, SIDS_LIST, ETF_LIST

# ── Monkeypatch: SPY filter on shorts only ──────────────────────────────────
_original_check_spy = backtest._check_spy_alignment

def _shorts_only_spy(spy_daily, date, signal_type):
    """Longs (OS) pass unconditionally; shorts still check SPY alignment."""
    if signal_type == "OS":
        return True
    return _original_check_spy(spy_daily, date, signal_type)

backtest._check_spy_alignment = _shorts_only_spy

# ── Style constants (matching output.py) ────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
RANK_FILLS = {
    "top25": PatternFill(start_color="548235", end_color="548235", fill_type="solid"),
    "top100": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "mid": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
    "bottom": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}
TOP25_FONT = Font(color="FFFFFF", size=11)
DEFAULT_FONT = Font(size=11)

CHECKPOINT_PATH = os.path.join(CACHE_DIR, "universe_checkpoint.json")

# ── Columns for ranked Excel ────────────────────────────────────────────────
RANKED_HEADERS = [
    "Rank", "★", "Ticker", "Sector", "Trades", "Wins", "Losses", "Win Rate",
    "Total P&L", "Avg P&L/Trade", "Avg RR", "Avg Duration",
    "Long Trades", "Long WR", "Short Trades", "Short WR",
    "RSI-50 Exit %", "Best Year WR", "Worst Year WR",
]


# ── Stat computation ────────────────────────────────────────────────────────

def compute_ticker_stats(ticker: str, trades: list[dict], sector: str,
                         is_sids: bool) -> dict:
    """Compute summary stats for one ticker's trades."""
    total = len(trades)
    if total == 0:
        return None

    wins = sum(1 for t in trades if t["win_loss"] == "Win")
    losses = total - wins
    win_rate = wins / total

    total_pnl = sum(t["total_profit"] for t in trades)
    avg_pnl = total_pnl / total
    avg_rr = sum(t["trade_rr"] for t in trades) / total
    avg_duration = sum(t["duration"] for t in trades) / total

    # Long/short breakdown
    longs = [t for t in trades if t["order"] == "Long"]
    shorts = [t for t in trades if t["order"] == "Short"]
    long_trades = len(longs)
    short_trades = len(shorts)
    long_wr = (sum(1 for t in longs if t["win_loss"] == "Win") / long_trades) if long_trades else 0
    short_wr = (sum(1 for t in shorts if t["win_loss"] == "Win") / short_trades) if short_trades else 0

    # Exit type breakdown from notes
    rsi50_exits = sum(1 for t in trades if "Exit:" not in t["notes"])
    stop_exits = sum(1 for t in trades if "Stop loss" in t.get("notes", ""))
    time_exits = sum(1 for t in trades if "Time exit" in t.get("notes", ""))
    rsi50_pct = rsi50_exits / total
    stop_pct = stop_exits / total
    time_pct = time_exits / total

    # Per-year breakdown
    year_trades = defaultdict(list)
    for t in trades:
        yr = t["entry_date"].year
        year_trades[yr].append(t)

    per_year = {}
    for yr, ytrades in sorted(year_trades.items()):
        yw = sum(1 for t in ytrades if t["win_loss"] == "Win")
        per_year[yr] = {
            "trades": len(ytrades),
            "wins": yw,
            "wr": yw / len(ytrades) if ytrades else 0,
        }

    year_wrs = [v["wr"] for v in per_year.values() if v["trades"] >= 3]
    best_yr_wr = max(year_wrs) if year_wrs else win_rate
    worst_yr_wr = min(year_wrs) if year_wrs else win_rate

    return {
        "ticker": ticker,
        "sector": sector,
        "is_sids": is_sids,
        "trades": total,
        "wins": wins,
        "losses": losses,
        "win_rate": win_rate,
        "total_pnl": round(total_pnl, 2),
        "avg_pnl": round(avg_pnl, 2),
        "avg_rr": round(avg_rr, 2),
        "avg_duration": round(avg_duration, 1),
        "long_trades": long_trades,
        "long_wr": round(long_wr, 4),
        "short_trades": short_trades,
        "short_wr": round(short_wr, 4),
        "rsi50_exit_pct": round(rsi50_pct, 4),
        "stop_pct": round(stop_pct, 4),
        "time_pct": round(time_pct, 4),
        "best_yr_wr": round(best_yr_wr, 4),
        "worst_yr_wr": round(worst_yr_wr, 4),
        "per_year": per_year,
    }


# ── Checkpoint system ───────────────────────────────────────────────────────

def _load_checkpoint() -> dict:
    """Load checkpoint if from today, else return empty."""
    if not os.path.exists(CHECKPOINT_PATH):
        return {}
    try:
        with open(CHECKPOINT_PATH, "r") as f:
            data = json.load(f)
        if data.get("date") != datetime.now().strftime("%Y-%m-%d"):
            return {}
        return data
    except Exception:
        return {}


def _save_checkpoint(completed: list[str], stats: dict[str, dict]):
    """Atomic save of checkpoint data."""
    os.makedirs(CACHE_DIR, exist_ok=True)
    data = {
        "date": datetime.now().strftime("%Y-%m-%d"),
        "completed": completed,
        "stats": stats,
    }
    # Atomic write: temp file + rename
    fd, tmp_path = tempfile.mkstemp(dir=CACHE_DIR, suffix=".json")
    try:
        with os.fdopen(fd, "w") as f:
            json.dump(data, f, default=str)
        os.replace(tmp_path, CHECKPOINT_PATH)
    except Exception:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise


# ── Excel generation ────────────────────────────────────────────────────────

def _write_ranked_sheet(ws, rows: list[dict], title: str = "Ranked"):
    """Write ranked ticker data to a worksheet."""
    # Header
    for col_idx, header in enumerate(RANKED_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"

    for row_idx, row in enumerate(rows, 2):
        rank = row_idx - 1
        values = [
            rank,
            "★" if row["is_sids"] else "",
            row["ticker"],
            row["sector"],
            row["trades"],
            row["wins"],
            row["losses"],
            row["win_rate"],
            row["total_pnl"],
            row["avg_pnl"],
            row["avg_rr"],
            row["avg_duration"],
            row["long_trades"],
            row["long_wr"],
            row["short_trades"],
            row["short_wr"],
            row["rsi50_exit_pct"],
            row["best_yr_wr"],
            row["worst_yr_wr"],
        ]

        # Determine fill by rank
        if rank <= 25:
            fill = RANK_FILLS["top25"]
            font = TOP25_FONT
        elif rank <= 100:
            fill = RANK_FILLS["top100"]
            font = DEFAULT_FONT
        elif rank <= 200:
            fill = RANK_FILLS["mid"]
            font = DEFAULT_FONT
        else:
            fill = RANK_FILLS["bottom"]
            font = DEFAULT_FONT

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.font = font

    # Number formatting
    last_row = len(rows) + 1
    for r in range(2, last_row + 1):
        ws.cell(row=r, column=8).number_format = '0.0%'    # Win Rate
        ws.cell(row=r, column=9).number_format = '$#,##0'   # Total P&L
        ws.cell(row=r, column=10).number_format = '$#,##0'  # Avg P&L/Trade
        ws.cell(row=r, column=11).number_format = '0.00'    # Avg RR
        ws.cell(row=r, column=12).number_format = '0.0'     # Avg Duration
        ws.cell(row=r, column=14).number_format = '0.0%'    # Long WR
        ws.cell(row=r, column=16).number_format = '0.0%'    # Short WR
        ws.cell(row=r, column=17).number_format = '0.0%'    # RSI-50 Exit %
        ws.cell(row=r, column=18).number_format = '0.0%'    # Best Year WR
        ws.cell(row=r, column=19).number_format = '0.0%'    # Worst Year WR

    # Auto-size columns
    for col_idx in range(1, len(RANKED_HEADERS) + 1):
        max_len = len(str(RANKED_HEADERS[col_idx - 1]))
        for r in range(2, min(last_row + 1, 52)):
            val = ws.cell(row=r, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 30)


def _write_summary_sheet(ws, qualifying: list[dict], non_qualifying: list[dict],
                         sector_map: dict, scan_start: datetime, scan_end: datetime,
                         total_scanned: int, errors: list[str]):
    """Write summary/metadata sheet."""
    bold = Font(bold=True, size=12)
    normal = Font(size=11)
    header_font = Font(bold=True, size=11)

    row = 1
    ws.cell(row=row, column=1, value="SID Universe Scan — Summary").font = Font(bold=True, size=14)
    row += 2

    # Scan metadata
    meta = [
        ("Scan Date", scan_start.strftime("%Y-%m-%d")),
        ("Scan Duration", f"{(scan_end - scan_start).total_seconds() / 60:.1f} minutes"),
        ("Total Tickers Scanned", total_scanned),
        ("Qualifying Tickers (>= 15 trades)", len(qualifying)),
        ("Non-Qualifying Tickers", len(non_qualifying)),
        ("Errors", len(errors)),
        ("SPY Filter", "Shorts only (longs pass unconditionally)"),
        ("Data Range", "2020-01-01 to present"),
    ]
    for label, val in meta:
        ws.cell(row=row, column=1, value=label).font = header_font
        ws.cell(row=row, column=2, value=val).font = normal
        row += 1

    row += 1

    # Win rate stats
    if qualifying:
        wrs = [q["win_rate"] for q in qualifying]
        ws.cell(row=row, column=1, value="Win Rate Statistics").font = bold
        row += 1
        wr_stats = [
            ("Average WR (qualifying)", f"{sum(wrs)/len(wrs):.1%}"),
            ("Median WR", f"{sorted(wrs)[len(wrs)//2]:.1%}"),
            ("Top 25 Avg WR", f"{sum(wrs[:25])/min(25,len(wrs)):.1%}"),
            ("Top 100 Avg WR", f"{sum(wrs[:100])/min(100,len(wrs)):.1%}"),
        ]
        for label, val in wr_stats:
            ws.cell(row=row, column=1, value=label).font = normal
            ws.cell(row=row, column=2, value=val).font = normal
            row += 1
        row += 1

    # Sid's list analysis
    sids_in_qual = [q for q in qualifying if q["is_sids"]]
    sids_top100 = [q for q in qualifying[:100] if q["is_sids"]]
    ws.cell(row=row, column=1, value="Sid's List Analysis").font = bold
    row += 1
    sid_stats = [
        ("Sid tickers qualifying", f"{len(sids_in_qual)} / {len(SIDS_LIST)}"),
        ("Sid tickers in Top 100", len(sids_top100)),
        ("Avg WR (Sid's list)", f"{sum(s['win_rate'] for s in sids_in_qual)/len(sids_in_qual):.1%}" if sids_in_qual else "N/A"),
    ]
    for label, val in sid_stats:
        ws.cell(row=row, column=1, value=label).font = normal
        ws.cell(row=row, column=2, value=val).font = normal
        row += 1
    row += 1

    # Sector breakdown
    ws.cell(row=row, column=1, value="Sector Breakdown").font = bold
    row += 1
    ws.cell(row=row, column=1, value="Sector").font = header_font
    ws.cell(row=row, column=2, value="Qualifying").font = header_font
    ws.cell(row=row, column=3, value="Avg WR").font = header_font
    row += 1

    sector_groups = defaultdict(list)
    for q in qualifying:
        sector_groups[q["sector"]].append(q["win_rate"])
    for sector, wrs in sorted(sector_groups.items(), key=lambda x: -sum(x[1])/len(x[1])):
        ws.cell(row=row, column=1, value=sector).font = normal
        ws.cell(row=row, column=2, value=len(wrs)).font = normal
        ws.cell(row=row, column=3, value=f"{sum(wrs)/len(wrs):.1%}").font = normal
        row += 1

    row += 2
    ws.cell(row=row, column=1,
            value="DISCLAIMER: This is in-sample backtesting data. Past performance "
                  "does not guarantee future results. Use as a starting point for "
                  "watchlist curation, not as a standalone trading signal.").font = Font(italic=True, size=10)

    # Auto-size
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15


def generate_ranked_excel(qualifying: list[dict]) -> str:
    """Generate the full ranked Excel file. Returns file path."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filename = f"sid_universe_ranked_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Universe Ranked"
    _write_ranked_sheet(ws, qualifying)
    wb.save(filepath)
    return filepath


def generate_top100_excel(qualifying: list[dict], non_qualifying: list[dict],
                          sector_map: dict, scan_start: datetime,
                          scan_end: datetime, total_scanned: int,
                          errors: list[str]) -> str:
    """Generate top-100 watchlist Excel with summary sheet. Returns file path."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filename = f"sid_top100_watchlist_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    wb = Workbook()

    # Sheet 1: Top 100
    ws1 = wb.active
    ws1.title = "Top 100"
    _write_ranked_sheet(ws1, qualifying[:100])

    # Sheet 2: Summary
    ws2 = wb.create_sheet("Summary")
    _write_summary_sheet(ws2, qualifying, non_qualifying, sector_map,
                         scan_start, scan_end, total_scanned, errors)

    wb.save(filepath)
    return filepath


# ── Terminal report ─────────────────────────────────────────────────────────

def print_terminal_report(qualifying: list[dict], non_qualifying: list[dict]):
    """Print summary report to terminal."""
    sids_set = set(SIDS_LIST)

    print(f"\n{'='*70}")
    print(f"  SID UNIVERSE SCAN — RESULTS")
    print(f"{'='*70}")

    # Top 20
    print(f"\n  TOP 20 BY WIN RATE")
    print(f"  {'Rank':<5} {'★':<2} {'Ticker':<8} {'Trades':>6} {'WR':>7} "
          f"{'P&L':>10} {'RR':>6} {'Sector'}")
    print(f"  {'-'*62}")
    for i, row in enumerate(qualifying[:20], 1):
        star = "★" if row["is_sids"] else " "
        print(f"  {i:<5} {star:<2} {row['ticker']:<8} {row['trades']:>6} "
              f"{row['win_rate']:>6.1%} {row['total_pnl']:>9,.0f} "
              f"{row['avg_rr']:>5.2f} {row['sector']}")

    # Bottom 5
    if len(qualifying) >= 5:
        print(f"\n  BOTTOM 5 QUALIFYING")
        print(f"  {'Rank':<5} {'★':<2} {'Ticker':<8} {'Trades':>6} {'WR':>7} "
              f"{'P&L':>10}")
        print(f"  {'-'*45}")
        for row in qualifying[-5:]:
            rank = qualifying.index(row) + 1
            star = "★" if row["is_sids"] else " "
            print(f"  {rank:<5} {star:<2} {row['ticker']:<8} {row['trades']:>6} "
                  f"{row['win_rate']:>6.1%} {row['total_pnl']:>9,.0f}")

    # Sid's list analysis
    sids_qual = [q for q in qualifying if q["is_sids"]]
    sids_top100 = [q for q in qualifying[:100] if q["is_sids"]]
    sids_under = [q for q in sids_qual if q["win_rate"] < 0.65]

    print(f"\n  SID'S LIST ANALYSIS")
    print(f"  Qualifying: {len(sids_qual)} / {len(SIDS_LIST)}")
    print(f"  In Top 100: {len(sids_top100)}")
    if sids_qual:
        avg_wr = sum(s["win_rate"] for s in sids_qual) / len(sids_qual)
        best = max(sids_qual, key=lambda x: x["win_rate"])
        worst = min(sids_qual, key=lambda x: x["win_rate"])
        print(f"  Avg WR: {avg_wr:.1%}")
        print(f"  Best: {best['ticker']} ({best['win_rate']:.1%})")
        print(f"  Worst: {worst['ticker']} ({worst['win_rate']:.1%})")
    if sids_under:
        print(f"  Underperformers (<65% WR): {', '.join(s['ticker'] for s in sids_under)}")

    # New discoveries
    new_top25 = [q for q in qualifying[:25] if not q["is_sids"]]
    if new_top25:
        print(f"\n  NEW DISCOVERIES (non-Sid in Top 25)")
        for q in new_top25:
            rank = qualifying.index(q) + 1
            print(f"    #{rank} {q['ticker']} — {q['win_rate']:.1%} WR, "
                  f"{q['trades']} trades ({q['sector']})")

    print(f"\n{'='*70}")


# ── Main scan ───────────────────────────────────────────────────────────────

def main():
    scan_start = datetime.now()

    # Build universe
    tickers, sector_map = get_universe()
    sids_set = set(SIDS_LIST)
    etf_set = set(ETF_LIST)
    total = len(tickers)

    print(f"\nStarting scan of {total} tickers...")
    print(f"SPY filter: shorts only (longs pass unconditionally)")
    print()

    # Pre-fetch SPY
    print("Loading SPY data for market alignment filter...", end=" ", flush=True)
    try:
        spy_daily = fetch_daily("SPY")
        if not spy_daily.empty:
            spy_daily = add_daily_indicators(spy_daily)
            print(f"{len(spy_daily)} candles loaded")
        else:
            print("WARNING: no SPY data — alignment filter disabled")
            spy_daily = None
    except Exception as e:
        print(f"WARNING: SPY fetch failed ({e}) — alignment filter disabled")
        spy_daily = None

    # Load checkpoint
    checkpoint = _load_checkpoint()
    completed = set(checkpoint.get("completed", []))
    all_stats = checkpoint.get("stats", {})
    if completed:
        print(f"Resuming from checkpoint: {len(completed)} tickers already done")

    errors = []
    processed_since_save = 0

    for i, ticker in enumerate(tickers, 1):
        # Skip SPY (it's the benchmark)
        if ticker == "SPY":
            completed.add(ticker)
            continue

        # Skip already completed
        if ticker in completed:
            continue

        print(f"[{i}/{total}] {ticker}...", end=" ", flush=True)

        try:
            # Fetch data
            daily_df = fetch_daily(ticker)
            if daily_df.empty:
                print("NO DATA — skipped")
                errors.append(f"{ticker}: no daily data")
                completed.add(ticker)
                processed_since_save += 1
                continue

            weekly_df = fetch_weekly(ticker)
            if weekly_df.empty:
                print("NO WEEKLY — skipped")
                errors.append(f"{ticker}: no weekly data")
                completed.add(ticker)
                processed_since_save += 1
                continue

            # Indicators
            daily_df = add_daily_indicators(daily_df)
            weekly_df = add_weekly_rsi(weekly_df)

            # Signals
            signals = find_rsi_signals(daily_df)

            # Earnings (skip for ETFs)
            if ticker in etf_set:
                earnings_dates = []
            else:
                earnings_dates = fetch_earnings_dates(ticker)

            # Backtest
            trades, skipped = run_backtest_for_ticker(
                ticker, daily_df, weekly_df, signals, earnings_dates, spy_daily
            )

            # Compute stats
            is_sids = ticker in sids_set
            sector = sector_map.get(ticker, "Other")
            stats = compute_ticker_stats(ticker, trades, sector, is_sids)

            if stats:
                all_stats[ticker] = stats
                qual_str = "QUAL" if stats["trades"] >= MIN_QUALIFYING_TRADES else "---"
                print(f"{len(signals)} sigs -> {stats['trades']} trades, "
                      f"WR {stats['win_rate']:.0%} [{qual_str}]")
            else:
                print("0 trades")

        except Exception as e:
            print(f"ERROR: {e}")
            errors.append(f"{ticker}: {e}")

        completed.add(ticker)
        processed_since_save += 1

        # Checkpoint every 50 tickers
        if processed_since_save >= 50:
            print(f"  [checkpoint: {len(completed)} tickers saved]")
            _save_checkpoint(list(completed), all_stats)
            processed_since_save = 0

        # Rate limiting
        time.sleep(0.5)

    # Final checkpoint
    _save_checkpoint(list(completed), all_stats)

    scan_end = datetime.now()

    # ── Rank and output ─────────────────────────────────────────────────
    # Split qualifying vs non-qualifying
    qualifying = []
    non_qualifying = []
    for stats in all_stats.values():
        if stats["trades"] >= MIN_QUALIFYING_TRADES:
            qualifying.append(stats)
        else:
            non_qualifying.append(stats)

    # Sort: WR desc, trade count desc (tiebreak)
    qualifying.sort(key=lambda x: (-x["win_rate"], -x["trades"]))

    print(f"\n{'='*60}")
    print(f"Scan complete: {len(all_stats)} tickers processed")
    print(f"Qualifying (>= {MIN_QUALIFYING_TRADES} trades): {len(qualifying)}")
    print(f"Non-qualifying: {len(non_qualifying)}")
    if errors:
        print(f"Errors: {len(errors)}")

    if qualifying:
        # Generate Excel files
        path1 = generate_ranked_excel(qualifying)
        print(f"\nRanked file: {path1}")

        path2 = generate_top100_excel(qualifying, non_qualifying, sector_map,
                                       scan_start, scan_end, total, errors)
        print(f"Top 100 file: {path2}")

        # Terminal report
        print_terminal_report(qualifying, non_qualifying)
    else:
        print("No qualifying tickers found.")

    if non_qualifying:
        print(f"\nNon-qualifying tickers ({len(non_qualifying)}):")
        for nq in sorted(non_qualifying, key=lambda x: -x["trades"])[:20]:
            star = "★" if nq["is_sids"] else " "
            print(f"  {star} {nq['ticker']}: {nq['trades']} trades, "
                  f"{nq['win_rate']:.0%} WR")
        if len(non_qualifying) > 20:
            print(f"  ... and {len(non_qualifying) - 20} more")


if __name__ == "__main__":
    main()
