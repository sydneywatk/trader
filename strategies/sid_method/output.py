"""Excel spreadsheet generation via openpyxl."""

import json
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR, CACHE_DIR


HEADERS = [
    "#",
    "TICKER",
    "Order",
    "Account Value",
    "Risk %",
    "$ Risk Per Position",
    "Stock Entry Price",
    "Stop Loss",
    "$ Risk Per Share",
    "% Risk Per Share",
    "Position Size $",
    "Max Shares",
    "RSI Signal Date",
    "Date Entered",
    "Date Exit",
    "Stock Exit Price",
    "$ Gain Per Share",
    "$ Total Profit",
    "Trade RR",
    "% Return",
    "Win/Loss",
    "Trade Duration",
    "Tier",
    "Suggested Action",
    "Notes",
]

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
WIN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
LOSS_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

TIER1_FILL = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
TIER1_FONT = Font(bold=True, color="FFFFFF", size=11)
TIER2_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
TIER3_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

TIER_ACTIONS = {
    1: "Options candidate OR 2-3% risk",
    2: "Standard trade at 1% risk",
    3: "Monitor only",
}


def _load_ticker_wr_map() -> dict[str, float]:
    """Load ticker historical WR from universe scan checkpoint."""
    checkpoint_path = os.path.join(CACHE_DIR, "universe_checkpoint.json")
    if not os.path.exists(checkpoint_path):
        return {}
    try:
        with open(checkpoint_path, "r") as f:
            data = json.load(f)
        return {t: s["win_rate"] for t, s in data.get("stats", {}).items()}
    except Exception:
        return {}


def score_trade_tier(trade: dict, ticker_wr: float | None) -> int:
    """Score a trade as Tier 1, 2, or 3 based on confluence of factors.

    Tier 1 — Options Candidate:
      Weekly RSI >5pts, MACD cross, RSI <25/<75, earnings >21d, ticker WR >80%
    Tier 2 — Strong Stock Trade:
      Weekly RSI 3-5pts, MACD pointing, RSI <30/>70, earnings >14d, ticker WR 70-80%
    Tier 3 — Monitor Only:
      Meets basic SID rules but not Tier 1 or 2.
    """
    signal_rsi = trade.get("signal_rsi", 50)
    weekly_delta = trade.get("weekly_rsi_delta", 0)
    macd_crossed = trade.get("macd_crossed", False)
    earnings_days = trade.get("earnings_days_away")
    order = trade.get("order", "Long")

    # Default WR to 0 if no data (will fail tier 1/2 WR check)
    wr = ticker_wr if ticker_wr is not None else 0

    # Tier 1 checks
    t1_weekly = weekly_delta > 5
    t1_macd = macd_crossed
    t1_rsi = (order == "Long" and signal_rsi < 25) or (order == "Short" and signal_rsi > 75)
    t1_earnings = earnings_days is not None and earnings_days > 21
    # If no earnings data, pass the earnings check (ETFs etc.)
    if earnings_days is None and not trade.get("notes", "").startswith(""):
        t1_earnings = True  # No earnings data = not a concern
    t1_wr = wr > 0.80

    t1_count = sum([t1_weekly, t1_macd, t1_rsi, t1_earnings, t1_wr])
    if t1_count >= 4:  # Need at least 4 of 5 criteria
        return 1

    # Tier 2 checks
    t2_weekly = weekly_delta >= 3
    t2_macd = True  # MACD is always pointing (it's an entry condition)
    t2_rsi = (order == "Long" and signal_rsi < 30) or (order == "Short" and signal_rsi > 70)
    t2_earnings = earnings_days is None or earnings_days > 14
    t2_wr = wr >= 0.70

    t2_count = sum([t2_weekly, t2_macd, t2_rsi, t2_earnings, t2_wr])
    if t2_count >= 4:  # Need at least 4 of 5 criteria
        return 2

    return 3


def generate_excel(trades: list[dict], skipped: list[dict]) -> str:
    """Generate the formatted Excel output. Returns the file path."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filename = f"sid_method_backtest_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "SID Backtest"

    # --- Header row ---
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Freeze top row
    ws.freeze_panes = "A2"

    # Sort trades by entry date
    trades_sorted = sorted(trades, key=lambda t: t["entry_date"])

    # Load ticker WR map for tier scoring
    wr_map = _load_ticker_wr_map()

    # Score tiers
    for trade in trades_sorted:
        trade["_tier"] = score_trade_tier(trade, wr_map.get(trade["ticker"]))

    # --- Data rows ---
    for row_idx, trade in enumerate(trades_sorted, 2):
        trade_num = row_idx - 1
        tier = trade["_tier"]

        values = [
            trade_num,
            trade["ticker"],
            trade["order"],
            trade["account_value"],
            trade["risk_pct"],
            trade["risk_per_position"],
            trade["entry_price"],
            trade["stop_loss"],
            trade["risk_per_share"],
            trade["pct_risk_per_share"],
            trade["position_size"],
            trade["max_shares"],
            trade["signal_date"].strftime("%m/%d/%Y"),
            trade["entry_date"].strftime("%m/%d/%Y"),
            trade["exit_date"].strftime("%m/%d/%Y"),
            trade["exit_price"],
            trade["gain_per_share"],
            trade["total_profit"],
            trade["trade_rr"],
            trade["pct_return"],
            trade["win_loss"],
            trade["duration"],
            tier,
            TIER_ACTIONS[tier],
            trade["notes"],
        ]

        fill = WIN_FILL if trade["win_loss"] == "Win" else LOSS_FILL
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill

        # Highlight tier column (col 23) with tier-specific styling
        tier_cell = ws.cell(row=row_idx, column=23)
        if tier == 1:
            tier_cell.fill = TIER1_FILL
            tier_cell.font = TIER1_FONT
        elif tier == 2:
            tier_cell.fill = TIER2_FILL

    # --- Number formatting ---
    last_data_row = len(trades_sorted) + 1
    for row_idx in range(2, last_data_row + 1):
        # Account Value (col 4)
        ws.cell(row=row_idx, column=4).number_format = '$#,##0'
        # Risk % (col 5)
        ws.cell(row=row_idx, column=5).number_format = '0.00%'
        # $ Risk Per Position (col 6)
        ws.cell(row=row_idx, column=6).number_format = '$#,##0.00'
        # Entry Price (col 7)
        ws.cell(row=row_idx, column=7).number_format = '$#,##0.00'
        # Stop Loss (col 8)
        ws.cell(row=row_idx, column=8).number_format = '$#,##0.00'
        # $ Risk Per Share (col 9)
        ws.cell(row=row_idx, column=9).number_format = '$#,##0.00'
        # % Risk Per Share (col 10)
        ws.cell(row=row_idx, column=10).number_format = '0.00%'
        # Position Size $ (col 11)
        ws.cell(row=row_idx, column=11).number_format = '$#,##0.00'
        # Max Shares (col 12)
        ws.cell(row=row_idx, column=12).number_format = '#,##0'
        # Exit Price (col 16)
        ws.cell(row=row_idx, column=16).number_format = '$#,##0.00'
        # $ Gain Per Share (col 17)
        ws.cell(row=row_idx, column=17).number_format = '$#,##0.00'
        # $ Total Profit (col 18)
        ws.cell(row=row_idx, column=18).number_format = '$#,##0.00'
        # Trade RR (col 19)
        ws.cell(row=row_idx, column=19).number_format = '0.00'
        # % Return (col 20)
        ws.cell(row=row_idx, column=20).number_format = '0.00%'
        # Trade Duration (col 22)
        ws.cell(row=row_idx, column=22).number_format = '#,##0'

    # --- Summary row ---
    summary_row = last_data_row + 2
    total_trades = len(trades_sorted)
    wins = sum(1 for t in trades_sorted if t["win_loss"] == "Win")
    losses = total_trades - wins
    win_rate = wins / total_trades if total_trades > 0 else 0
    total_pnl = sum(t["total_profit"] for t in trades_sorted)
    avg_rr = sum(t["trade_rr"] for t in trades_sorted) / total_trades if total_trades > 0 else 0

    summary_font = Font(bold=True, size=11)

    # Per-tier stats
    tier_stats = {}
    for tier_num in (1, 2, 3):
        tier_trades = [t for t in trades_sorted if t["_tier"] == tier_num]
        t_count = len(tier_trades)
        t_wins = sum(1 for t in tier_trades if t["win_loss"] == "Win")
        t_wr = t_wins / t_count if t_count else 0
        t_pnl = sum(t["total_profit"] for t in tier_trades)
        t_avg_pnl = t_pnl / t_count if t_count else 0
        tier_stats[tier_num] = {"count": t_count, "wins": t_wins, "wr": t_wr,
                                "pnl": t_pnl, "avg_pnl": t_avg_pnl}

    summaries = [
        (1, "SUMMARY"),
        (2, f"Total Trades: {total_trades}"),
        (3, f"Wins: {wins}"),
        (4, f"Losses: {losses}"),
        (5, f"Win Rate: {win_rate:.1%}"),
        (6, f"Total P&L: ${total_pnl:,.2f}"),
        (7, f"Avg Trade RR: {avg_rr:.2f}"),
    ]
    for col, val in summaries:
        cell = ws.cell(row=summary_row, column=col, value=val)
        cell.font = summary_font

    # Tier breakdown in summary
    tier_row = summary_row + 2
    ws.cell(row=tier_row, column=1, value="TIER BREAKDOWN").font = summary_font
    tier_row += 1
    for tier_num in (1, 2, 3):
        ts = tier_stats[tier_num]
        label = {1: "Tier 1 (Options)", 2: "Tier 2 (Strong)", 3: "Tier 3 (Monitor)"}[tier_num]
        ws.cell(row=tier_row, column=1, value=label).font = summary_font
        ws.cell(row=tier_row, column=2, value=f"Trades: {ts['count']}")
        ws.cell(row=tier_row, column=3, value=f"WR: {ts['wr']:.1%}")
        ws.cell(row=tier_row, column=4, value=f"P&L: ${ts['pnl']:,.2f}")
        ws.cell(row=tier_row, column=5, value=f"Avg P&L: ${ts['avg_pnl']:,.2f}")
        tier_row += 1

    # --- Auto-size columns ---
    for col_idx in range(1, len(HEADERS) + 1):
        max_len = len(str(HEADERS[col_idx - 1]))
        for row_idx in range(2, min(last_data_row + 1, 52)):  # Sample first 50 rows
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        adjusted_width = min(max_len + 3, 45)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # --- Skipped trades sheet ---
    if skipped:
        ws2 = wb.create_sheet("Skipped Trades")
        skip_headers = ["Ticker", "Signal Date", "Type", "Reason"]
        for col_idx, h in enumerate(skip_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        for row_idx, s in enumerate(skipped, 2):
            ws2.cell(row=row_idx, column=1, value=s["ticker"])
            ws2.cell(row=row_idx, column=2, value=s["signal_date"].strftime("%m/%d/%Y"))
            ws2.cell(row=row_idx, column=3, value=s["signal_type"])
            ws2.cell(row=row_idx, column=4, value=s["reason"])
        for col_idx in range(1, 5):
            ws2.column_dimensions[get_column_letter(col_idx)].width = 30
        ws2.freeze_panes = "A2"

    wb.save(filepath)
    return filepath
