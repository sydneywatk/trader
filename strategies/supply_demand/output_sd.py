"""Excel output for Supply & Demand backtest results."""

from __future__ import annotations

import os
from datetime import datetime
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR

# Distinct color palette from SID (SID uses dark blue headers)
HEADER_FILL = PatternFill(start_color="1E5631", end_color="1E5631", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WIN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
LOSS_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HIGH_PRIORITY_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

HEADERS = [
    "ticker",
    "direction",
    "zone_type",
    "priority",
    "confirmation_candle",
    "signal_date",
    "entry_date",
    "entry_price",
    "stop_loss",
    "take_profit",
    "exit_date",
    "exit_price",
    "exit_reason",
    "gain_loss_dollars",
    "gain_loss_pct",
    "trade_rr",
    "trade_duration",
    "win_loss",
    "risk_dollars",
    "shares",
    "atr_at_entry",
    "htf_aligned",
    "zone_age_at_entry",
    "zone_proximal",
    "zone_distal",
    "earnings_days_away",
]


def _fmt_date(v) -> str:
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    return str(v)


def _write_trades_sheet(ws, trades: list[dict]) -> None:
    # Header
    for col, name in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    # Rows
    for r, tr in enumerate(trades, start=2):
        fill = WIN_FILL if tr["win_loss"] == "Win" else LOSS_FILL
        for col, name in enumerate(HEADERS, start=1):
            val = tr.get(name, "")
            if name in ("signal_date", "entry_date", "exit_date"):
                val = _fmt_date(val)
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = fill
            # Highlight zone_type cell for high priority
            if name == "zone_type" and tr.get("priority") == "high":
                cell.fill = HIGH_PRIORITY_FILL
                cell.font = Font(bold=True)

    # Summary row
    if trades:
        sum_row = len(trades) + 3
        wins = sum(1 for t in trades if t["win_loss"] == "Win")
        losses = len(trades) - wins
        wr = wins / len(trades) * 100.0
        total_pl = sum(t["gain_loss_dollars"] for t in trades)
        avg_rr = sum(t["trade_rr"] for t in trades) / len(trades)
        avg_dur = sum(t["trade_duration"] for t in trades) / len(trades)

        summary = [
            ("Total trades", len(trades)),
            ("Wins", wins),
            ("Losses", losses),
            ("Win Rate", f"{wr:.1f}%"),
            ("Total P&L", f"${total_pl:,.2f}"),
            ("Avg RR", f"{avg_rr:.2f}"),
            ("Avg Duration", f"{avg_dur:.1f}d"),
        ]
        for i, (label, value) in enumerate(summary):
            lc = ws.cell(row=sum_row, column=1 + i * 2, value=label)
            lc.font = Font(bold=True)
            ws.cell(row=sum_row, column=2 + i * 2, value=value)

    # Auto-size (rough): use max content length per column
    for col in range(1, len(HEADERS) + 1):
        letter = get_column_letter(col)
        max_len = len(HEADERS[col - 1])
        for r in range(2, len(trades) + 2):
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_len + 2, 32)


def _write_by_zone_type(ws, trades: list[dict]) -> None:
    ws.freeze_panes = "A2"
    headers = ["zone_type", "priority", "trades", "wins", "win_rate", "avg_pl", "avg_rr", "avg_duration"]
    for col, name in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    buckets: dict[str, list[dict]] = {"DBR": [], "RBD": [], "RBR": [], "DBD": []}
    for t in trades:
        buckets.setdefault(t["zone_type"], []).append(t)

    r = 2
    for zt in ("DBR", "RBD", "RBR", "DBD"):
        bucket = buckets.get(zt, [])
        if not bucket:
            ws.cell(row=r, column=1, value=zt)
            ws.cell(row=r, column=2, value="high" if zt in ("DBR", "RBD") else "low")
            ws.cell(row=r, column=3, value=0)
            r += 1
            continue
        wins = sum(1 for t in bucket if t["win_loss"] == "Win")
        wr = wins / len(bucket) * 100.0
        avg_pl = sum(t["gain_loss_dollars"] for t in bucket) / len(bucket)
        avg_rr = sum(t["trade_rr"] for t in bucket) / len(bucket)
        avg_dur = sum(t["trade_duration"] for t in bucket) / len(bucket)
        ws.cell(row=r, column=1, value=zt)
        ws.cell(row=r, column=2, value="high" if zt in ("DBR", "RBD") else "low")
        ws.cell(row=r, column=3, value=len(bucket))
        ws.cell(row=r, column=4, value=wins)
        ws.cell(row=r, column=5, value=f"{wr:.1f}%")
        ws.cell(row=r, column=6, value=f"${avg_pl:,.2f}")
        ws.cell(row=r, column=7, value=f"{avg_rr:.2f}")
        ws.cell(row=r, column=8, value=f"{avg_dur:.1f}")
        if zt in ("DBR", "RBD"):
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).fill = HIGH_PRIORITY_FILL
        r += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14


def _write_by_ticker(ws, trades: list[dict], min_trades: int = 5) -> None:
    ws.freeze_panes = "A2"
    headers = ["ticker", "trades", "wins", "win_rate", "total_pl", "avg_rr"]
    for col, name in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    by_ticker: dict[str, list[dict]] = {}
    for t in trades:
        by_ticker.setdefault(t["ticker"], []).append(t)

    rows = []
    for tk, ts in by_ticker.items():
        if len(ts) < min_trades:
            continue
        wins = sum(1 for t in ts if t["win_loss"] == "Win")
        wr = wins / len(ts) * 100.0
        total = sum(t["gain_loss_dollars"] for t in ts)
        avg_rr = sum(t["trade_rr"] for t in ts) / len(ts)
        rows.append((tk, len(ts), wins, wr, total, avg_rr))

    rows.sort(key=lambda x: x[3], reverse=True)
    top_20 = rows[:20]
    bottom_10 = rows[-10:] if len(rows) >= 10 else []

    r = 2
    def _write_group(title: str, group: list[tuple]) -> int:
        nonlocal r
        c = ws.cell(row=r, column=1, value=title)
        c.font = Font(bold=True)
        r += 1
        for tk, n, w, wr, pl, arr in group:
            ws.cell(row=r, column=1, value=tk)
            ws.cell(row=r, column=2, value=n)
            ws.cell(row=r, column=3, value=w)
            ws.cell(row=r, column=4, value=f"{wr:.1f}%")
            ws.cell(row=r, column=5, value=f"${pl:,.2f}")
            ws.cell(row=r, column=6, value=f"{arr:.2f}")
            r += 1
        r += 1
        return r

    _write_group(f"Top 20 (min {min_trades} trades)", top_20)
    if bottom_10:
        _write_group("Bottom 10", bottom_10)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14


def generate_excel(trades: list[dict], output_dir: str = OUTPUT_DIR) -> str:
    """Write sd_method_backtest_YYYYMMDD.xlsx. Returns the absolute path."""
    os.makedirs(output_dir, exist_ok=True)
    fname = f"sd_method_backtest_{datetime.now().strftime('%Y%m%d')}.xlsx"
    path = os.path.join(output_dir, fname)

    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "All Trades"
    _write_trades_sheet(ws_all, trades)

    ws_zt = wb.create_sheet("By Zone Type")
    _write_by_zone_type(ws_zt, trades)

    ws_tk = wb.create_sheet("By Ticker")
    _write_by_ticker(ws_tk, trades)

    wb.save(path)
    return path
